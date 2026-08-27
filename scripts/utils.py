"""
共享工具模块——北京时间、周末守卫、Excel 读写、日志。

核心教训：沙盒时钟是 UTC，所有日期判断必须显式转北京时间。
"""
import os
import json
import logging
from datetime import datetime, timedelta, timezone

import openpyxl

# ==================== 常量 ====================

BEIJING_TZ = timezone(timedelta(hours=8))
UTC_TZ = timezone.utc

MAGNITUDE_SMALL = 0.3   # |涨跌幅| <= 0.3% 为小
MAGNITUDE_MEDIUM = 1.0  # |涨跌幅| <= 1.0% 为中
MAGNITUDE_LARGE = 1.0    # |涨跌幅| > 1.0% 为大

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, "data")
LOG_PATH = os.path.join(REPO_ROOT, "log.xlsx")
DASHBOARD_PATH = os.path.join(REPO_ROOT, "dashboard", "dashboard.html")

SHEET_LOG = "预测日志"
SHEET_STATS = "统计"
SHEET_RULES = "说明"

# 预测日志列映射 (1-based index for openpyxl)
COL = {
    "target_date": 1,      # A 目标日期
    "gen_time": 2,         # B 生成时间
    "base_price": 3,       # C 基准价($/oz)
    "final_dir": 4,        # D 最终方向
    "final_mag": 5,        # E 最终幅度档
    "model_dir": 6,        # F 模型原始方向
    "model_mag": 7,        # G 模型原始档位
    "human_adj": 8,        # H 人工修正
    "actual_price": 9,     # I 次日实际价
    "actual_change": 10,   # J 实际涨跌幅
    "actual_dir": 11,      # K 实际方向
    "actual_mag": 12,      # L 实际档位
    "dir_correct": 13,     # M 方向对错
    "mag_correct": 14,     # N 档位对错
    "market": 15,          # O 市况标签
    "drivers": 16,         # P 关键驱动摘要
    "note": 17,            # Q 备注
    "confidence": 18,      # R 置信度(%)
}

# ==================== 日志 ====================

def get_logger(name: str) -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S"
        ))
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
    return logger

# ==================== 北京时间 ====================

def now_beijing() -> datetime:
    """返回当前北京时间（显式从 UTC 转换，不依赖系统时区）。"""
    return datetime.now(UTC_TZ).astimezone(BEIJING_TZ)

def beijing_date_str() -> str:
    """返回北京日期字符串 YYYY-MM-DD。"""
    return now_beijing().strftime("%Y-%m-%d")

def beijing_datetime_str() -> str:
    """返回北京时间字符串 YYYY-MM-DD HH:MM。"""
    return now_beijing().strftime("%Y-%m-%d %H:%M")

def to_beijing_weekday() -> int:
    """返回北京时间的星期几（0=周一, 6=周日）。"""
    return now_beijing().weekday()

def is_weekend() -> bool:
    """周末休市守卫——北京时间周六、周日不交易。"""
    return to_beijing_weekday() >= 5

def yesterday_beijing_date_str() -> str:
    """返回昨天（北京时间）的日期字符串。"""
    return (now_beijing() - timedelta(days=1)).strftime("%Y-%m-%d")

# ==================== 幅度档判定 ====================

def get_magnitude(change_pct: float) -> str:
    """根据涨跌幅百分比返回幅度档：小/中/大。"""
    abs_chg = abs(change_pct)
    if abs_chg <= MAGNITUDE_SMALL:
        return "小"
    elif abs_chg <= MAGNITUDE_MEDIUM:
        return "中"
    else:
        return "大"

def get_direction(change_pct: float) -> str:
    """根据涨跌幅返回方向：涨/跌/平。"""
    if change_pct > 0:
        return "涨"
    elif change_pct < 0:
        return "跌"
    else:
        return "平"

# ==================== Excel 读写 ====================

def load_workbook():
    """加载 log.xlsx，返回 (workbook, log_sheet)。"""
    wb = openpyxl.load_workbook(LOG_PATH)
    ws = wb[SHEET_LOG]
    return wb, ws

def find_row_by_date(ws, target_date: str) -> int:
    """在预测日志中按目标日期查找行号，找不到返回 -1。"""
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=COL["target_date"]).value
        if val and str(val).strip() == target_date.strip():
            return row
    return -1

def find_last_data_row(ws) -> int:
    """找到最后一行有数据的行号。"""
    last = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=COL["target_date"]).value:
            last = row
    return last

def save_workbook(wb):
    """保存 workbook。"""
    wb.save(LOG_PATH)

def get_all_records(ws) -> list:
    """读取所有记录，返回字典列表。"""
    records = []
    for row in range(2, ws.max_row + 1):
        target_date = ws.cell(row=row, column=COL["target_date"]).value
        if not target_date:
            continue

        def read_cell(col):
            """读取单元格值，公式单元格返回空字符串。"""
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, str) and val.startswith("="):
                return ""  # 公式视为未填写
            return val

        records.append({
            "row": row,
            "target_date": str(target_date),
            "gen_time": str(read_cell(COL["gen_time"]) or ""),
            "base_price": read_cell(COL["base_price"]),
            "final_dir": read_cell(COL["final_dir"]) or "",
            "final_mag": read_cell(COL["final_mag"]) or "",
            "model_dir": read_cell(COL["model_dir"]) or "",
            "model_mag": read_cell(COL["model_mag"]) or "",
            "human_adj": read_cell(COL["human_adj"]) or "",
            "actual_price": read_cell(COL["actual_price"]),
            "actual_change": read_cell(COL["actual_change"]),
            "actual_dir": read_cell(COL["actual_dir"]) or "",
            "actual_mag": read_cell(COL["actual_mag"]) or "",
            "dir_correct": read_cell(COL["dir_correct"]) or "",
            "mag_correct": read_cell(COL["mag_correct"]) or "",
            "market": read_cell(COL["market"]) or "",
            "drivers": read_cell(COL["drivers"]) or "",
            "note": read_cell(COL["note"]) or "",
            "confidence": read_cell(COL["confidence"]),
        })
    return records

def update_stats_sheet(wb):
    """根据日志数据更新统计 Sheet。"""
    ws_log = wb[SHEET_LOG]
    ws_stats = wb[SHEET_STATS]
    records = get_all_records(ws_log)

    settled = [r for r in records if r["dir_correct"] in ("√", "×")]
    total = len(records)
    dir_correct = sum(1 for r in settled if r["dir_correct"] == "√")
    mag_correct = sum(1 for r in settled if r["mag_correct"] == "√")

    settled_count = len(settled)
    dir_rate = round(dir_correct / settled_count * 100, 1) if settled_count else None
    mag_rate = round(mag_correct / settled_count * 100, 1) if settled_count else None

    # 近 30 次
    recent_settled = settled[-30:]
    recent_count = len(recent_settled)
    recent_dir = sum(1 for r in recent_settled if r["dir_correct"] == "√")
    recent_mag = sum(1 for r in recent_settled if r["mag_correct"] == "√")
    recent_dir_rate = round(recent_dir / recent_count * 100, 1) if recent_count else None
    recent_mag_rate = round(recent_mag / recent_count * 100, 1) if recent_count else None

    # 分市况
    trend_settled = [r for r in settled if "趋势" in (r["market"] or "")]
    range_settled = [r for r in settled if "震荡" in (r["market"] or "")]
    trend_dir = sum(1 for r in trend_settled if r["dir_correct"] == "√")
    range_dir = sum(1 for r in range_settled if r["dir_correct"] == "√")
    trend_rate = round(trend_dir / len(trend_settled) * 100, 1) if trend_settled else None
    range_rate = round(range_dir / len(range_settled) * 100, 1) if range_settled else None

    # 置信度统计
    confidence_recorded = [r for r in records if r["confidence"] is not None]
    confidence_settled = [r for r in settled if r["confidence"] is not None]

    stats_map = {
        2: settled_count,
        3: total,
        4: dir_rate,
        5: mag_rate,
        6: recent_dir_rate,
        7: recent_mag_rate,
        8: trend_rate,
        9: range_rate,
        12: len(confidence_recorded),
        13: len(confidence_settled),
    }
    for row_num, value in stats_map.items():
        ws_stats.cell(row=row_num, column=2, value=value if value is not None else "-")

# ==================== 数据文件读写 ====================

def save_data(filename: str, data):
    """保存数据到 data/ 目录下的 JSON 文件。"""
    os.makedirs(DATA_DIR, exist_ok=True)
    path = os.path.join(DATA_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path

def load_data(filename: str):
    """从 data/ 目录读取 JSON 文件。"""
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# ==================== 市况判定 ====================

def determine_market_label(price_history: list) -> str:
    """
    根据近 20 日价格序列判定市况标签。
    返回：趋势涨 / 趋势跌 / 震荡
    """
    if len(price_history) < 2:
        return "震荡"
    recent = price_history[-20:] if len(price_history) >= 20 else price_history
    first = recent[0]["usd"]
    last = recent[-1]["usd"]
    change_pct = (last - first) / first * 100

    if change_pct > 3:
        return "趋势涨"
    elif change_pct < -3:
        return "趋势跌"
    else:
        return "震荡"
